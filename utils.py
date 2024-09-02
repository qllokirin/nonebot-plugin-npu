import imgkit
import os
import json
from openpyxl.styles import Border, Side, Alignment, Font, PatternFill, Alignment

# 未完成课程为红色填充
red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
# 未完成模块为黄色填充
yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
# 未匹配课程为灰色填充
gray_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

# Function to generate HTML table from data
def generate_html_table(data):
    html_table = """
    <style>
        table {
            width: 100%;
            border-collapse: collapse;
        }
        th, td {
            border: 1px solid #dddddd;
            text-align: left;
            padding: 8px;
        }
        th {
            background-color: #f2f2f2;
        }
        tr:nth-child(even) {
            background-color: #f9f9f9;
        }
    </style>
    <table>
        <thead>
            <tr>
                <th>Name</th>
                <th>Grade Score</th>
                <th>GPA</th>
                <th>Credit</th>
                <th>Grade Detail</th>
            </tr>
        </thead>
        <tbody>
    """

    for item in data:
        html_table += "<tr>"
        for value in item.values():
            if isinstance(value, list):
                html_table += f"<td>{'<br>'.join(value)}</td>"
            else:
                html_table += f"<td>{value}</td>"
        html_table += "</tr>"

    html_table += """
        </tbody>
    </table>
    """

    return html_table


def generate_img_from_html(data, grades_folder_path):
    # Generate HTML table
    html_table_content = generate_html_table(data)

    # Save HTML content to a file
    with open(os.path.join(grades_folder_path, "grades.html"), "w", encoding="utf-8") as html_file:
        html_file.write(html_table_content)

    # Convert HTML to image using imgkit
    options = {
        'format': 'jpg',
        'encoding': 'utf-8',
    }

    imgkit.from_file(os.path.join(grades_folder_path, "grades.html"), os.path.join(grades_folder_path, "grades.jpg"),
                     options=options)


def generate_grades_to_msg(gardes):
    grades_msg = ""
    for garde in gardes:
        grades_msg += "名称：" + garde["name"] + "\n"
        grades_msg += "分数：" + garde["grade_score"] + "\n"
        grades_msg += "绩点：" + garde["gpa"] + "\n"
        grades_msg += "学分：" + str(garde["credit"]) + "\n"
        if garde["grade_detail"]:
            grades_msg += "详细成绩:" + "\n"
            for detail in garde["grade_detail"][:-1]:
                grades_msg += "├──" + detail + "\n"
            grades_msg += "└──" + garde["grade_detail"][-1] + "\n\n"
    return grades_msg[:-2]

def get_exams_msg(folder_path):
    with open((os.path.join(folder_path, 'exams.json')), 'r', encoding='utf-8') as f:
        exams = json.loads(f.read())
    exams_msg = ""
    for exam in exams:
        exams_msg += "名称：" + exam["course"] +"\n"
        exams_msg += "地点：" + exam["location"] +"\n"
        exams_msg += "时间：" + exam["time"] +"\n\n"
    exams_msg = exams_msg[:-2]
    return exams_msg

# 处理成绩数据
def handle_training_program_data(data, results):
    for item in data:
        remark = item.get("remark", "")
        class_info = {
            "type_nameZh": item["type"]["nameZh"],
            "remark": remark,
            "requiredCredits": item["requireInfo"]["requiredCredits"],
            "planCourses": []
        }
        data_plan = item.get("planCourses", [])
        if data_plan:
            for data_one in data_plan:
                course_info = {
                    "course_nameZh": data_one["course"]["nameZh"],
                    "course_code": data_one["course"]["code"],
                    "course_credits": data_one["course"]["credits"],
                    "course_type": data_one["course"]["courseType"]["nameZh"]
                }
                class_info["planCourses"].append(course_info)
        else:
            children = item.get("children", [])
            if children:
                class_info["children"] = []
                handle_training_program_data(children, class_info["children"])
        results.append(class_info)

# 递归正确计算剩余未修学分
def calculate_remaining_credits(data):
    if "children" in data:
        # 遍历所有直接子节点，递归计算并累加其 remainingCredits
        for child in data["children"]:
            # 递归计算子节点的 remainingCredits
            calculate_remaining_credits(child)
        child_remaining_credits = [child["remainingCredits"] for child in data["children"]]
        child_required_credits = [child["requiredCredits"] for child in data["children"]]
        # 当学分要求大于子节点的总学分时，说明是分组的培养方案，取子节点中最小的 remainingCredits
        if sum(child_required_credits) > data["requiredCredits"]:
            data["remainingCredits"] = min(child_remaining_credits)
            data["children"] = [child for child in data["children"] if child["remainingCredits"] == data["remainingCredits"]]
        else:
            data["remainingCredits"] = sum(child_remaining_credits)
        if data["remainingCredits"] < 0:
            data["remainingCredits"] = 0

# 处理匹配和未匹配的课程
def handle_completed_and_incomplete_course(program, completed_courses_all, completed_courses_all_static):
    for item in program:
        # 初始化已修学分和未修学分
        required_credits = item.get("requiredCredits", 0)
        completed_credits = 0
        incomplete_courses = []
        completed_courses = []
        # 处理课程计划中的每一门课程
        for course in item.get("planCourses", []):
            course_code = course["course_code"]
            course_credits = course.get("course_credits", 0)
            if course_code in completed_courses_all_static:
                completed_credits += course_credits
                completed_courses.append(course)
                course["completed"] = True
                if completed_courses_all.get(course_code):
                    del completed_courses_all[course_code]
            else:
                incomplete_courses.append(course)
                course["completed"] = False
        item["planCourses"] = [] 
        item["completedCredits"] = completed_credits
        item["remainingCredits"] = required_credits - completed_credits
        item["incompleteCourses"] = incomplete_courses
        item["completedCourses"] = completed_courses
        if item["remainingCredits"] < 0:
            item["remainingCredits"] = 0
        # 递归处理子分类
        if "children" in item:
            handle_completed_and_incomplete_course(item["children"], completed_courses_all, completed_courses_all_static)
        if "children" not in item and item["incompleteCourses"] == []:
            item["incompleteCourses"].append({"course_nameZh": "培养方案无具体课程", "course_code": "无", "course_credits": "无", "course_type": "无"})
        # 递归正确计算剩余未修学分
        calculate_remaining_credits(item)

# 删除 requiredCredits 为 0 的分组
def del_zero_required_credits_group(nodes):
    filtered_nodes = []
    for node in nodes:
        if "children" in node:
            node["children"] = del_zero_required_credits_group(node["children"])
        if node["requiredCredits"] != 0:
            filtered_nodes.append(node)
    return filtered_nodes

# 计算最大层级深度
def max_dict_depth(data):
    if isinstance(data, dict):
        # 如果是字典，递归计算每个子项的最大深度
        return 1 + max((max_dict_depth(v) for v in data.values()), default=0)
    elif isinstance(data, list):
        # 如果是列表，递归计算每个元素的最大深度
        return max((max_dict_depth(item) for item in data), default=0)
    else:
        # 如果是其他类型，深度为0
        return 0

# 写入数据到Excel表格
def write_to_excel(data, sheet, max_depth, row=1, col=1):
    for item in data:
        # 写入当前层级的名称# 确定是否显示备注部分
        remark_text = f"\n备注:{item['remark']}" if item['remark'] else ""
        # 创建单元格并动态设置内容
        cell = sheet.cell(
            row=row,
            column=col,
            value=f"{item['type_nameZh']}\n需修{item['requiredCredits']}学分\n还剩{item['remainingCredits']}学分{remark_text}"
        )
        cell.alignment = Alignment(wrap_text=True)
        # 如果 remainingCredits 大于 0，设置单元格填充颜色为黄色
        if item['remainingCredits'] > 0:
            cell.fill = yellow_fill
        # 如果有子节点，递归写入
        if "children" in item and item["children"]:
            # 递归处理子节点
            row = write_to_excel(item["children"], sheet, max_depth, row, col+1)
        # 写入课程信息
        if "incompleteCourses" in item and item["remainingCredits"] > 0:
            for course in item["incompleteCourses"]:
                sheet.cell(row=row, column=col + 1, value=course["course_nameZh"]).fill = red_fill
                sheet.cell(row=row, column=col + 2, value=course["course_code"]).fill = red_fill
                sheet.cell(row=row, column=col + 3, value=course["course_type"]).fill = red_fill
                sheet.cell(row=row, column=col + 4, value=course["course_credits"]).fill = red_fill
                for col_temp in range(col+5, max_depth + 4):
                    sheet.cell(row=row, column=col_temp, value='空')
                row += 1
        if "completedCourses" in item:
            for course in item["completedCourses"]:
                sheet.cell(row=row, column=col+1, value=course["course_nameZh"]).font = Font(color="0000FF")
                sheet.cell(row=row, column=col+2, value=course["course_code"]).font = Font(color="0000FF")
                sheet.cell(row=row, column=col+3, value=course["course_type"]).font = Font(color="0000FF")
                sheet.cell(row=row, column=col+4, value=course["course_credits"]).font = Font(color="0000FF")
                for col_temp in range(col+5, max_depth + 4):
                    sheet.cell(row=row, column=col_temp, value='空')
                row += 1
        completed_courses = item.get("completedCourses")
        incomplete_courses = item.get("incompleteCourses")
        children = item.get("children")
        if (completed_courses == []) and (incomplete_courses == []) and (children is None):
            row += 1
    return row

# 格式化表格
def fromat_excel(sheet, completed_courses_all):
    # 遍历所有单元格
    for col in sheet.columns:
        col_letter = col[0].column_letter
        max_row = sheet.max_row
        row = 1
        while row <= max_row:
            current_cell = sheet[f"{col_letter}{row}"]
            # 如果当前单元格为空，则跳到下一行
            if not current_cell.value:
                row += 1
                continue
            # 查找下方连续空单元格
            merge_start = row
            while row + 1 <= max_row and not sheet[f"{col_letter}{row + 1}"].value:
                row += 1
            # 如果发现需要合并的范围，进行合并
            if merge_start != row:
                sheet.merge_cells(start_row=merge_start, start_column=current_cell.column, end_row=row, end_column=current_cell.column)
            # 跳到下一个未合并的单元格
            row += 1
    for row in sheet.iter_rows():
        for cell in row:
            # 如果单元格的值为字符 '空'，则清空单元格
            if cell.value == '空':
                cell.value = None
    # 假设开始填充数据的起始行
    start_row = sheet.max_row + 1
    # 将未匹配的课程并填充到Excel表格
    for i, (_, course) in enumerate(completed_courses_all.items(), start=start_row):
        # 填充课程信息到指定列
        sheet.cell(row=i, column=1, value=course["name"]).fill = gray_fill
        sheet.cell(row=i, column=2, value=course["code"]).fill = gray_fill
        sheet.cell(row=i, column=3, value=course["course_type"]).fill = gray_fill
        sheet.cell(row=i, column=4, value=course["credit"]).fill = gray_fill
    # 设置框线（上下左右） 垂直居中并自动换行
    for row in sheet.iter_rows():
        for cell in row:
            cell.border = Border(
                            left=Side(border_style="thin", color="000000"),
                            right=Side(border_style="thin", color="000000"),
                            top=Side(border_style="thin", color="000000"),
                            bottom=Side(border_style="thin", color="000000")
                            )
            cell.alignment = Alignment(vertical="center", wrap_text=True)
    # 设置所有列的宽度为20
    for col in sheet.columns:
        col_letter = col[0].column_letter
        sheet.column_dimensions[col_letter].width = 20
